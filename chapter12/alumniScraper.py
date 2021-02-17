#! /usr/bin/env bash
# python3 alumniScraper.py - Scrapes LSE alumni emails from the US alumni website, for the sake of practice only 

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import pyinputplus, sys, time, logging, pprint, re, csv, openpyxl
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format=' %(asctime)s -  %(levelname)s -  %(message)s')
# logging.disable(logging.CRITICAL)

# get email and password
userEmail = sys.argv[1]
userPass = sys.argv[2]

# quickLoad function for practicing in shell
def quickLoad(mail, passwrd):
    userEmail = mail
    userPass = passwrd
    # open chrome browser
    browser = webdriver.Chrome()
    # go to alumni site login page
    browser.get('http://www.aflse.org/user.html?op=login')
    # enter email
    browser.find_element_by_name('email').send_keys(userEmail)
    # enter password
    browser.find_element_by_name('password').send_keys(userPass)
    # login
    browser.find_element_by_name('op').click()
    time.sleep(2)
    # go to alumni directory
    directory = browser.find_element_by_link_text('US Alumni Directory')
    directory.click()
    time.sleep(2)
    # filter by city
    browser.find_element_by_name('filter_location_city').send_keys('New York')
    # search for results
    browser.find_element_by_name('submit').click()
    time.sleep(2)
    return browser

browser = quickLoad(userEmail, userPass)

# get alum cards and names to interate over

alumCardElems = []
alumNameElems = []

# Scroll down page to load all results
# Get scroll height
lastHeight = browser.execute_script("return document.body.scrollHeight;")
time.sleep(2)
while True:
    # scroll to end
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    # scroll up by two x height of screen because when scrolling to bottom sometimes it doesn't load
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight - 2 * screen.height);")
    time.sleep(.5)
    # scroll down by one x height of screen to start loading if possible
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight - 1 * screen.height);")
    time.sleep(1.5)
    # check if anything has loaded
    if browser.execute_script("return document.body.scrollHeight;") == lastHeight: # i.e. nothing new loaded, then stop scrolling
        break
    # else update last height and keep scrolling down to continue loading
    lastHeight = browser.execute_script("return document.body.scrollHeight;")

        
# get results of alumni cards
alumCardElems = alumCardElems + browser.find_elements_by_css_selector('.row.bizcard')
alumNameElems = alumNameElems + browser.find_elements_by_css_selector('.media-heading.bizcard_zoom_activator')
logging.info('length of alumCardElems: ' + str(len(alumCardElems)) + '| length of alumNameElems: ' + str(len(alumNameElems)))

alumList = []

# get alum Ids
alumIds = []
for alum in alumCardElems:
    alumIds.append(alum.get_attribute('id'))

# for each .row.bizcard alumCardElem
for index, nameElem in enumerate(alumNameElems):
    logging.info(f'index: {index}')
    alum = {}
    # get alums Id
    alumId = str(alumIds[index])
    # reveal alum info
    nameElem.click()
    time.sleep(2)
    # Get alum name
    alum["name"] = nameElem.text
    # Get location
    # set xpath for addresses using alums Id
    locationPathX = "//*[@id='" + alumId + "']//div[@class='col-sm-10']"
    locationWithName = browser.find_element_by_xpath(locationPathX).text # contains alum name
    locationNoName = locationWithName.replace(alum["name"], "").strip()
    alum["location"] = locationNoName
    # Get addresses
    # set xpath for addresses using alums Id
    #homePathX = "//div[@id='" + alumId + "']//a[@class='btn btn-primary btn-xs']"
    addressPathX = "//*[@id='postcard_" + alumId + "']/div[2]"
    alum["addresses"] = browser.find_element_by_xpath(addressPathX).text
    # Get work experience
    experiencePathX = "//div[@id='" + alumId + "']//b"
    # check if have work experience, if not set to null
    try:
        alum["experience"] = browser.find_element_by_xpath(experiencePathX).text
    except:
        alum["experience"] = None
    # Get emails
    # check if have primary email, if not set to null    
    # set xpath for primary email using alums Id
    primaryEmailPathX = "//div[@id='" + alumId + "']//a[@class='btn btn-primary btn-xs']"
    try:
        emailElem = browser.find_element_by_xpath(primaryEmailPathX)
        alum["primaryEmail"] = emailElem.get_attribute('href')[7:]  # splice to remove the 'mailto:' at beginning of email
    except:
        alum["primaryEmail"] = None
    # check if have a secondary email, if not set to null
    # set xpath for secondary email using alums Id
    secondaryEmailPathX = "//div[@id='" + alumId + "']//a[@class='btn btn-info btn-xs']"
    try:
        secEmailElem = browser.find_element_by_xpath(secondaryEmailPathX)
        alum["secondaryEmail"] = secEmailElem.get_attribute('href')[7:] # splice to remove the 'mailto:' at beginning of email
    except:
        alum["secondaryEmail"] = None
    # Add alum to alum list 
    alumList.append(alum)
    logging.debug(f"{alum}")

# Parse addresses in alumList into separate work and home addresses

def countNewLines(addresses):
    newLineRegex = re.compile("\n")
    mo = newLineRegex.findall(addresses)
    return len(mo)

# set regex to detect if have both addresses, or just one of work or home address
regexBoth = re.compile(r'((work: |Work: |Seasonal: )(.*)Home: (.*))', re.DOTALL)
regexWork = re.compile(r'((work: |Work: |Seasonal: )(.*))', re.DOTALL)
regexHome = re.compile(r'(Home: (.*))', re.DOTALL)

print(f'alumList[-2:]:')
pprint.pprint(alumList[-2:])


for alum in alumList:
    numberLines = countNewLines(alum['addresses'])
    # if both work and home addresses
    try:
        if numberLines == 2: 
            moBoth = regexBoth.search(alum['addresses'])
            alum["workAddress"] = moBoth.group(3)
            alum["homeAddress"] = moBoth.group(4)
        # if only one address, check what type of address
        elif numberLines == 1:
            moWork = regexWork.search(alum['addresses'])
            moHome = regexHome.search(alum['addresses'])
            # check if work address
            if moWork != None:
                alum["workAddress"] = moWork.group(3)
                alum["homeAddress"] = None
            else:
                alum["homeAddress"] = moHome.group(2)
                alum["workAddress"] = None
        # if no addresses
        else:
            alum["homeAddress"] = None
            alum["workAddress"] = None
    except: # in case of weird address formats
        alum["homeAddress"] = None
        alum["workAddress"] = None
        alum.pop('addresses', None) 
        logging.debug(f'alum: {alum} | alum[addresses]: {alum["addresses"]}')
        continue
    # delete addresses key from alum
    alum.pop('addresses', None) 

# pprint.pprint(alumList)

# write to a .py file
fileObj = open('/Users/laurencefinch/Desktop/AutomateBoringStuff/alumList3.py', 'w')
fileObj.write('alums = ' + pprint.pformat(alumList) + '\n')
fileObj.close()

# write to a CSV file
outputFile = open('/Users/laurencefinch/Desktop/AutomateBoringStuff/alumList3.csv', 'w')
outputDictWriter = csv.DictWriter(outputFile, ['name', 'location', 'experience', 'primaryEmail', 'secondaryEmail', 'workAddress', 'homeAddress'])
outputDictWriter.writeheader()
for alum in alumList:
    outputDictWriter.writerow(alum)
outputFile.close()

# write to Excel file 
outputExcel = openpyxl.Workbook()
sheet = outputExcel.active
sheet.title = "New York alums"
# set Headers
i = 0
for header in ['name', 'location', 'experience', 'homeAddress', 'workAddress', 'primaryEmail', 'secondaryEmail']:
    sheet[get_column_letter(i) + str(1)] = header
    i += 1
# loop over alums
currentRow = 1
for alum in alumList:
    currentRow += 1
    # add alum details to the corresponding for that row (alum)
    for col in range(1, 8):
        sheet.cell(row=currentRow, column=col).value = alum[sheet.cell(row=1, column=col).value]
# save Excel
outputExcel.save('/Users/laurencefinch/Desktop/AutomateBoringStuff/alumList.xlsx')
