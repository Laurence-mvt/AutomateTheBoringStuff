# csvToExcel - adds alumni csv files for different locations to excel file as different sheets (for educational purposes only)

import csv
import openpyxl, logging, datetime
from pathlib import Path
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG, format=' \n%(asctime)s -  %(levelname)s -  %(message)s\n')
logging.disable(logging.CRITICAL)

# get csv files paths from directory
p = Path('/Users/laurencefinch/Desktop/AutomateBoringStuff/alumScrapeResults')
csvFilePaths = list(p.glob('*.csv'))

# open Excel workbook
wb = openpyxl.Workbook()
sheet = wb.active

# loop over csv files
for csvFilePath in csvFilePaths:

    logging.info(f'Now working on csvFile {csvFilePath}')
    
    # open csv file and get location from name and alums in csvDictReader
    csvFile = open(csvFilePath)
    location = csvFilePath.stem
    alumReader = csv.DictReader(csvFile)
    alumList = list(alumReader)
    csvFile.close()

    # create new sheet for that location
    wb.create_sheet(title=location)
    
    # work on that sheet
    sheet = wb[location]

    # add headers
    headers = alumReader.fieldnames
    for i in range(1, len(headers)+1):
        sheet[f'{get_column_letter(i)}1'] = headers[i-1]

    # add alumni data
    # loop over all alums
    rowNum = 2  # first alum's row, after headers
    for i, alum in enumerate(alumList[len(alumList) - 2:]):
        logging.debug(f'{alum}')
        # loop over fields for that alum's row
        for index, header in enumerate(headers):
            sheet[f'{get_column_letter(index+1)}{rowNum}'] = alumList[len(alumList) - 2 + i][header]
        rowNum += 1

        rowNum = 2  # first alum's row, after headers
    for i, alum in enumerate(alumList[1:]):
        logging.debug(f'{alum}')
        # loop over fields for that alum's row
        for index, header in enumerate(headers):
            sheet[f'{get_column_letter(index+1)}{rowNum}'] = alumList[i][header]
        rowNum += 1
    # add last alum
    for index, header in enumerate(headers):
        sheet[f'{get_column_letter(index+1)}{rowNum}'] = alumList[-1][header]
    
# save workbook
# get datetime
updateTime = datetime.datetime.now()
# format to ddMonyyyy
updateTime = updateTime.strftime('%d%b%Y')
wb.save(f'/Users/laurencefinch/Desktop/AutomateBoringStuff/alumScrapeResults/alumList4_{updateTime}.xlsx')
