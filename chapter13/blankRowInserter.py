# blankRowInserter.py - inserts user-inputted number of blank rows at a user-inputted row in Excel
# note: this program returns two new Exels. 
# One where it does not update formulas, when shifted down by M lines, so if a cell sums cols B:D on that row, after shifting down it will still sum cols B:D for that initial pre-shift row.
# Another where it updates the formula ('=ROUND(B[row]*C[row], 2)') in the fourth column after shifting rows (not in the Automate BoringStuff book)

import openpyxl, sys

# check user inputted two numbers
if len(sys.argv) != 4:
    print('Usage: python3 blankRowInserter.py N M exelFile.xlsx   - where N (first row to insert into) are integers and M (number of lines to insert)')
    exit()

N = int(sys.argv[1]) # row to insert first blank row
M = int(sys.argv[2]) # number of blank rows to insert

# for without formulas updating
# open blank workbook and get sheet
wbNew = openpyxl.Workbook()
newSheet = wbNew.active

# for with formulas updating in fourth column
# open blank workbook and get sheet
wbNewFormulas = openpyxl.Workbook()
newSheetFormulas = wbNewFormulas.active

# open existing workbook and get sheet
wbOld = openpyxl.load_workbook(sys.argv[3])
oldSheet = wbOld.active

# copy first N - 1 rows 
for i in range(1, N):
    # loop over each cell in that row
    for col in range(1, oldSheet.max_column + 1):
        newSheet.cell(row=i, column=col).value = oldSheet.cell(row=i, column=col).value # first row on new sheet = first row on old sheet.. etc.
        newSheetFormulas.cell(row=i, column=col).value = oldSheet.cell(row=i, column=col).value

numCols = oldSheet.max_column

# from row M in new sheet, add M to row index
for i in range(N, oldSheet.max_row + 1):
    # loop over each cell in that row - FOR WITHOUT FORMULAS
    for col in range(1, numCols + 1):
        # cell value for new sheet is same as cell value for old sheet, but M rows below
        newSheet.cell(row=i+M, column=col).value = oldSheet.cell(row=i, column=col).value
    # loop over each cell in that row - FOR WITH FORMULAS IN FINAL COLUMN
    for col in range(1, numCols): # ignore final column
        # cell value for new sheet is same as cell value for old sheet, but M rows below
        newSheetFormulas.cell(row=i+M, column=col).value = oldSheet.cell(row=i, column=col).value
    # for final column where formula is '=ROUND(B[row]*C[row], 2)'
    newSheetFormulas.cell(row=i+M, column = numCols).value = f'=ROUND(B{i+M}*C{i+M}, 2)'
    
# save new workbooks
wbNew.save(f'new{sys.argv[3]}')
wbNewFormulas.save(f'FormulasNew{sys.argv[3]}')
