# multiplicationTable.py - creates an NxN multiplication table in Excel for user inputed N

import openpyxl, sys

if len(sys.argv) != 2:
    print('Usage: python3 multiplicationTable.py N  where N is the NxN size of desired table')
    exit()

# get size of table, N
N = int(sys.argv[1])

# create workbook 
wb = openpyxl.Workbook()
sheet = wb.active

# add numbers up to N in top row and first column
for i in range(2, N+2):
    # top row
    sheet.cell(row=1, column=i).value = i - 1
    # first column
    sheet.cell(row=i, column=1).value = i - 1

# loop over rows and perform calculations
for row in range(2, N + 2):
    # loop over each column
    for col in range(2, N + 2):
        # set values as product of row value and col value
        sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=1).value * sheet.cell(row=1, column=col).value

# save workbook
wb.save('multiplicationTable.xlsx') 