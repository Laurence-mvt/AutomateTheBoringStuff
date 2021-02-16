#! usr/bin/env bash python3 

# This book uses version 2.6.2 of OpenPyXL. 
# It’s important that you install this version by running pip install --user -U openpyxl==2.6.2 
# because newer versions of OpenPyXL are incompatible with the information in this book. 

import openpyxl

wb = openpyxl.load_workbook('example.xlsx') 
type(wb)    # <class 'openpyxl.workbook.workbook.Workbook'>

# get sheetnames and sheet properties
wb.sheetnames   # ['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb['Sheet1']
sheet.title 
# get the active sheet
wb.active

# get cells in sheets
sheet['A1']    # for cell A1
sheet['A1'].value  # 2015-04-05 13:34:02 automatatically interpreted as a datetime value
b = sheet['A1']
b.value     # 'Apples'
print(f'Row {b.row}, Column {b.column} is {b.value}')   # Row 1, Column 1 is 2015-04-05 13:34:02
print(f'Cell {b.coordinate} is {b.value}')  # Cell A1 is 2015-04-05 13:34:02

for i in range(1, 8, 2): # Go through every other row:
    print(i, sheet.cell(row=i, column=2).value)

# get size of sheet in rows and columns
sheet.max_row
sheet.max_column

# translate column numbers to letters using openpyxl.utils
openpyxl.utils.get_column_letter(1)     # 'A'
openpyxl.utils.get_column_letter(900)   # 'AHP'
from openpyxl.utils import get_column_letter, column_index_from_string
column_index_from_string('A')   # 1

# splice a worksheet to get a specific range of cells
tuple(sheet['A1':'C3'])    # ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

# get columsn or rows
row2 = list(sheet.rows)[1] # to get second row (as 0 indexed list). note sheet.rows (and .columns) returns a tuple of tuples, one tuple for each row/ col containing each cell as items  
column2 = list(sheet.columns)[1] # to get second column

for cellObj in row2:
    print(cellObj.value)

# Writing Excel docs
wb = openpyxl.Workbook() # Create a blank workbook
wb.sheetnames # ['Sheet']
sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet' # Change sheet title
wb.save('example_copy.xlsx') # Save the workbook as example_copy.xlsx
# best to always save an existing workbook as a copy with a different name in case you get a bug in your code that screws things up

# add a sheet
wb.create_sheet(index=0, title='FirstSheet') # creates a new sheet at index 0, retuns that sheet object
wb.sheetnames # ['FirstSheet', 'Spam Bacon Eggs Sheet']
wb.create_sheet(title='NewSheet')
del wb['NewSheet']  # delete a worksheet
wb.save('example_copy.xlsx')

# Writing values to cells
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'hello, world!' # Edit the cell's value
sheet['A1'].value   # 'hello, world!'

# Style cells
# to set font styles
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True) # Create a font
sheet['A1'].font = italic24Font # apply the font to A1
sheet['A1'] = 'hello,world!'
wb.save('styles.xlsx')

# Formulas
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)' # Set the formula
wb.save('writeFormula.xlsx')

# Adjust Rows and Columns
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
# set the height and width:
sheet.row_dimensions[1].height = 70 
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

# merge cells
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3') # merge all these cells
sheet['A1'] = 'Twelve cells merged together'
sheet.merge_cells('C5:D5') # merge these two cells
sheet['C5'] = 'Two merged cells'
wb.save('merged.xlsx')

# unmerge 
sheet.unmerge_cells('A1:D3') # unmerge these cells
wb.save('merged.xlss')

# freeze panes
sheet.freeze_panes = 'A2' # freeze first row
sheet.freeze_panes = 'C2' # freeze first row and columns A and B
sheet.freeze_panes = None

# Charts
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11): # create some data in column !
    sheet['A'+str(i)] = i

refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)   # create a reference object
seriesObj = openpyxl.chart.Series(refObj, title='First series') # create a series object from refernce object
chartObj = openpyxl.chart.BarChart()    # create  a bar chart
chartObj.title = 'My Chart'
chartObj.append(seriesObj)

sheet.add_chart(chartObj, 'C5') # add chart to sheet with top left corner at cell C5
wb.save('sampleChart.xlsx')

# can also create line charts, scatter charts, and pie charts

